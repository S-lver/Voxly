import os
from openpyxl import load_workbook
from openpyxl import Workbook
from models import Student, db

class ExcelService:
    def __init__(self, upload_folder='uploads'):
        self.upload_folder = upload_folder
        if not os.path.exists(upload_folder):
            os.makedirs(upload_folder)
    
    def save_file(self, file):
        filename = file.filename
        filepath = os.path.join(self.upload_folder, filename)
        file.save(filepath)
        return filepath
    
    def process_students(self, filepath):
        try:
            wb = load_workbook(filepath, data_only=True)
            ws = wb.active
            
            headers = []
            for cell in ws[1]:
                if cell.value:
                    headers.append(str(cell.value).strip().lower())
                else:
                    headers.append('')
            
            col_map = {}
            for idx, header in enumerate(headers):
                if 'name' in header:
                    col_map['name'] = idx
                elif 'grade' in header:
                    col_map['grade'] = idx
                elif 'homeroom' in header or 'room' in header:
                    col_map['homeroom'] = idx
                elif 'phone' in header or 'parent' in header:
                    col_map['parent_phone'] = idx
                elif 'balance' in header:
                    col_map['balance'] = idx
                elif 'attendance' in header:
                    col_map['attendance'] = idx
                elif 'email' in header:
                    col_map['email'] = idx
            
            if 'name' not in col_map:
                return {'error': 'Column "Name" not found in Excel file'}
            
            students_added = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[col_map.get('name', 0)]:
                    continue
                
                student = Student(
                    name=str(row[col_map.get('name', 0)]).strip(),
                    grade=str(row[col_map.get('grade', 0)]).strip() if col_map.get('grade') is not None and row[col_map.get('grade', 0)] else '',
                    homeroom=str(row[col_map.get('homeroom', 0)]).strip() if col_map.get('homeroom') is not None and row[col_map.get('homeroom', 0)] else '',
                    parent_phone=str(row[col_map.get('parent_phone', 0)]).strip() if col_map.get('parent_phone') is not None and row[col_map.get('parent_phone', 0)] else '',
                    balance=float(row[col_map.get('balance', 0)]) if col_map.get('balance') is not None and row[col_map.get('balance', 0)] and row[col_map.get('balance', 0)] is not None else 0.0,
                    attendance=str(row[col_map.get('attendance', 0)]).strip() if col_map.get('attendance') is not None and row[col_map.get('attendance', 0)] else 'Present',
                    email=str(row[col_map.get('email', 0)]).strip() if col_map.get('email') is not None and row[col_map.get('email', 0)] else ''
                )
                db.session.add(student)
                students_added += 1
            
            db.session.commit()
            return {'success': True, 'students_added': students_added}
            
        except Exception as e:
            return {'error': str(e)}
    
    def export_sample(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Students"
        
        headers = ['Name', 'Grade', 'Homeroom', 'ParentPhone', 'Balance', 'Attendance', 'Email']
        ws.append(headers)
        
        sample_data = [
            ['Mike Johnson', 10, 'A101', '+1234567890', 45.50, 'Present', 'mike.j@email.com'],
            ['Sarah Chen', 12, 'B203', '+1987654321', 0.00, 'Present', 'sarah.c@email.com'],
            ['James Rodriguez', 11, 'C305', '+1555123456', 12.75, 'Absent', 'james.r@email.com'],
        ]
        for row in sample_data:
            ws.append(row)
        
        filepath = os.path.join(self.upload_folder, 'sample_students.xlsx')
        wb.save(filepath)
        return filepath